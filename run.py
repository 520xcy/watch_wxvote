# coding=utf-8

import time
import shelve
import os
import sys
import signal
import openpyxl
import tkinter as tk
from tkinter import filedialog

from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait

BASE_PATH = os.getcwd()
DATA_PATH = BASE_PATH + "/votedata"
# 微信投票的地址
URL = 'https://mp.weixin.qq.com/s/THi2tZEqI8zPlAUxTpZlZA'
# 显示多少次取值
COUNT = 10
# 刷新时间
WAITTIME = 60


def save_file():
    window = tk.Tk()
    window.withdraw()
    file_path = filedialog.asksaveasfilename(title=u'保存文件', filetypes=[(
        "xlsx", ".xlsx"), ('All Files', '*')], initialdir=(os.path.expanduser(BASE_PATH)))
    file_path = file_path+'.xlsx'
    return file_path


def pushData(data):
    with shelve.open(DATA_PATH) as write:
        write[time.strftime('%Y-%m-%d %H:%M:%S',
                            time.localtime(time.time()))] = data


def getData(count=0):
    data = {}
    with shelve.open(DATA_PATH) as read:
        keys = list(read.keys())
        try:
            keys.sort()
        except:
            pass
        if not count == 0:
            keys = keys[-count:]
        for key in keys:
            data[key] = read[key]
    return data


def getVote():
    # 投票的微信页面地址
    BROWER.get(URL)
    # 投票的iframe的class
    voteiframe = BROWER.find_elements_by_class_name('js_editor_vote_card')[0]
    # 获取iframe的实际地址
    data_src = voteiframe.get_attribute('data-src')

    # 访问实际的投票地址
    BROWER.get("https://mp.weixin.qq.com" + data_src)
    # 返回投票数据变量名（具体可在页面源码中找到）
    voteInfo = BROWER.execute_script('return voteInfo')
    return voteInfo


def run():
    votes = getVote()
    votes = votes['vote_subject'][0]['options']
    pushData(votes)
    votes = getData(COUNT)
    data = {}
    title = []
    for votetime in votes:
        title.append(votetime.split(' '))
        for vote in votes[votetime]:
            if vote['name'] in data:
                data[vote['name']] = str(
                    data[vote['name']]) + '|' + str(vote['cnt']).rjust(10, ' ')
            else:
                data[vote['name']] = str(vote['cnt']).rjust(10, ' ')

    keys = data.keys()
    keyL = len(max(data.keys(), key=len))
    re = {}
    for key in keys:
        alpha_num = 0
        for i in range(len(key)):
            if ord(key[i]) in range(65, 91) or ord(key[i]) in range(97, 123) or ord(key[i]) == 32:
                alpha_num += 1
        re[key.ljust(keyL+int(alpha_num/2), '　')] = data[key]
    title_date = '　'*keyL+' '
    title_time = '　'*keyL+' '
    for key in title:
        title_date += str(key[0])+'|'
        title_time += str(key[1]).center(10, ' ')+'|'

    print(title_date)
    print(title_time)

    for key in re:
        print(key+':'+re[key]+'|')


def quit(signum, frame):
    BROWER.quit()
    print('正在退出...')
    exit()


def export():

    votes = getData()
    date = list(votes.keys())

    data = {}
    for votetime in votes:
        for vote in votes[votetime]:
            if not vote['name'] in data:
                data[vote['name']] = []
            data[vote['name']].append(vote['cnt'])

    f = openpyxl.Workbook()  # 创建工作簿

    sheet1 = f.create_sheet(title='votes')

    names = list(data.keys())
    for d in range(len(date)):
        sheet1.cell(row=1, column=d+2).value = str(date[d])
    for r in range(len(names)):
        sheet1.cell(row=r+2, column=1).value = str(names[r])
        for c in range(len(data[names[r]])):
            sheet1.cell(row=r+2, column=c+2).value = int(data[names[r]][c])

    f.save(save_file())  # 保存文件


if __name__ == "__main__":
    if len(sys.argv) > 1:
        if sys.argv[1] == 'export':
            export()
        else:
            print('未知参数')
    else:
        chromeOptions = webdriver.ChromeOptions()
        chromeOptions.add_argument('--headless')
        BROWER = webdriver.Chrome(options=chromeOptions)
        wait = WebDriverWait(BROWER, 10)
        BROWER.maximize_window()
        BROWER.implicitly_wait(6)
        signal.signal(signal.SIGINT, quit)
        signal.signal(signal.SIGTERM, quit)
        while True:
            os.system('clear')
            run()
            time.sleep(WAITTIME)
